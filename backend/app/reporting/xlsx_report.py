import io

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models import ScanResult, ScanRun
from app.reporting.summary import compute_scan_summary

GREEN = "10B981"
AMBER = "F59E0B"
RED = "EF4444"
GRAY = "94A3B8"
HEADER_BG = "1E293B"

STATUS_FILL = {
    "ONLINE": PatternFill("solid", fgColor=GREEN),
    "WARNING": PatternFill("solid", fgColor=AMBER),
    "OFFLINE": PatternFill("solid", fgColor=RED),
}
PROTECTED_FILL = PatternFill("solid", fgColor=GREEN)
UNPROTECTED_FILL = PatternFill("solid", fgColor=RED)
EXTERNAL_FILL = PatternFill("solid", fgColor=RED)
INTERNAL_FILL = PatternFill("solid", fgColor=GRAY)
HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
HEADER_FONT = Font(color="FFFFFF", bold=True)

S3_STATUS_LABELS = {
    "public": "PÚBLICO",
    "private": "Privado",
    "not_found": "Não encontrado",
    "unknown": "Indeterminado",
}
S3_SOURCE_LABELS = {
    "cname_direct": "CNAME direto (confirmado)",
    "cname_cdn": "Atrás de CDN/WAF (não confirmado)",
    "guess": "Mesmo nome do domínio (não confirmado)",
}
S3_PUBLIC_FILL = PatternFill("solid", fgColor=RED)


def _exposure_label(r: ScanResult) -> str:
    if r.is_internal is True:
        return "Interno"
    if r.is_internal is False:
        return "Externo"
    return "Desconhecido"


def _exposure_fill(r: ScanResult) -> PatternFill | None:
    if r.is_internal is True:
        return INTERNAL_FILL
    if r.is_internal is False:
        return EXTERNAL_FILL
    return None


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")


def _autosize_columns(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build_summary_sheet(wb: Workbook, scan_run: ScanRun, summary: dict, project_name: str) -> None:
    ws = wb.active
    ws.title = "Resumo"

    ws["A1"] = f"Relatório Executivo — Cobertura WAF (Akamai) — {project_name}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Scan run #{scan_run.id}  ·  {scan_run.started_at:%d/%m/%Y %H:%M}"
    ws["A2"].font = Font(italic=True, color="64748B")

    metrics = [
        ("Total de domínios", summary["total"]),
        ("Online", summary["online"]),
        ("Warning (HTTP != 200)", summary["warning"]),
        ("Offline", summary["offline"]),
        ("% Online", f"{summary['pct_online']}%"),
        ("Protegidos por Akamai", summary["protected"]),
        ("% Cobertura WAF", f"{summary['pct_protected']}%"),
        ("Domínios expostos externamente", summary["external_count"]),
        ("Domínios apenas internos (IP privado)", summary["internal_count"]),
        ("Exposição desconhecida (DNS não resolvido)", summary["unknown_exposure_count"]),
        ("EXTERNO, online e sem proteção (risco real)", summary["external_unprotected_online"]),
        ("Interno, online e sem proteção (risco baixo)", summary["internal_unprotected_online"]),
        ("Buckets S3 verificados", summary["s3_checked_count"]),
        ("BUCKETS S3 PÚBLICOS (risco crítico)", summary["s3_public_count"]),
    ]

    start_row = 4
    ws.cell(row=start_row, column=1, value="Métrica")
    ws.cell(row=start_row, column=2, value="Valor")
    _style_header_row(ws, start_row, 2)

    for i, (label, value) in enumerate(metrics, start=start_row + 1):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2, value=value)
        if label.startswith("EXTERNO") or label.startswith("BUCKETS S3 PÚBLICOS"):
            cell.fill = UNPROTECTED_FILL
            cell.font = Font(bold=True, color="FFFFFF")
        elif label.startswith("Interno, online"):
            cell.fill = INTERNAL_FILL
            cell.font = Font(bold=True, color="FFFFFF")

    _autosize_columns(ws, [46, 14])

    chart_start = start_row + len(metrics) + 3
    ws.cell(row=chart_start, column=1, value="Categoria")
    ws.cell(row=chart_start, column=2, value="Domínios")
    ws.cell(row=chart_start + 1, column=1, value="Protegido (Akamai)")
    ws.cell(row=chart_start + 1, column=2, value=summary["protected"])
    ws.cell(row=chart_start + 2, column=1, value="Externo e sem proteção (risco)")
    ws.cell(row=chart_start + 2, column=2, value=summary["external_unprotected_online"])
    ws.cell(row=chart_start + 3, column=1, value="Interno e sem proteção")
    ws.cell(row=chart_start + 3, column=2, value=summary["internal_unprotected_online"])
    ws.cell(row=chart_start + 4, column=1, value="Offline e sem proteção")
    ws.cell(row=chart_start + 4, column=2, value=summary["unprotected_offline"])

    chart = PieChart()
    chart.title = "Cobertura WAF x Exposição"
    data = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + 4)
    categories = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + 4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 13
    ws.add_chart(chart, f"D{start_row}")


def _build_risk_sheet(wb: Workbook, results: list[ScanResult]) -> None:
    ws = wb.create_sheet("Risco - Sem Proteção")
    headers = ["Hostname", "Exposição", "URL Verificada", "Status", "Código HTTP", "IP Resolvido", "Verificado em"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    risky = [r for r in results if r.status in ("ONLINE", "WARNING") and not r.akamai_protected]
    # Externo primeiro (risco real), depois interno, depois exposição desconhecida
    exposure_order = {False: 0, True: 1, None: 2}
    risky.sort(key=lambda r: (exposure_order[r.is_internal], r.domain.hostname))

    for r in risky:
        row = [
            r.domain.hostname,
            _exposure_label(r),
            r.checked_url,
            r.status,
            r.http_status_code,
            r.resolved_ip,
            r.checked_at.strftime("%d/%m/%Y %H:%M"),
        ]
        ws.append(row)
        line = ws.max_row
        ws.cell(row=line, column=4).fill = STATUS_FILL[r.status]
        exposure_fill = _exposure_fill(r)
        if exposure_fill:
            ws.cell(row=line, column=2).fill = exposure_fill
            ws.cell(row=line, column=2).font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws, [32, 14, 40, 12, 12, 16, 18])


def _build_s3_sheet(wb: Workbook, results: list[ScanResult]) -> None:
    ws = wb.create_sheet("Buckets S3")
    headers = ["Hostname", "Nome do Bucket", "Status", "Confiança", "Sinais", "Verificado em"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    checked = [r for r in results if r.s3_status is not None]
    status_order = {"public": 0, "unknown": 1, "private": 2, "not_found": 3}
    checked.sort(key=lambda r: (status_order.get(r.s3_status, 9), r.domain.hostname))

    for r in checked:
        row = [
            r.domain.hostname,
            r.s3_bucket_name,
            S3_STATUS_LABELS.get(r.s3_status, r.s3_status),
            S3_SOURCE_LABELS.get(r.s3_source, r.s3_source),
            "; ".join(r.s3_signals or []),
            r.checked_at.strftime("%d/%m/%Y %H:%M"),
        ]
        ws.append(row)
        if r.s3_status == "public":
            line = ws.max_row
            ws.cell(row=line, column=3).fill = S3_PUBLIC_FILL
            ws.cell(row=line, column=3).font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    if checked:
        ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws, [32, 32, 16, 32, 50, 18])


def _build_detail_sheet(wb: Workbook, results: list[ScanResult]) -> None:
    ws = wb.create_sheet("Detalhado")
    headers = [
        "Hostname", "Exposição", "URL Verificada", "Status", "Código HTTP", "Erro",
        "Tempo Resposta (ms)", "IP Resolvido", "Protegido Akamai", "Sinais Akamai",
        "Cadeia CNAME", "TLS Issuer", "TLS Expira em", "S3 Status", "S3 Bucket", "Verificado em",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for r in sorted(results, key=lambda r: r.domain.hostname):
        row = [
            r.domain.hostname,
            _exposure_label(r),
            r.checked_url,
            r.status,
            r.http_status_code,
            r.error_message,
            round(r.response_time_ms, 1) if r.response_time_ms else None,
            r.resolved_ip,
            "Sim" if r.akamai_protected else "Não",
            "; ".join(r.akamai_signals),
            " -> ".join(r.cname_chain),
            r.tls_issuer,
            r.tls_expiry.strftime("%d/%m/%Y") if r.tls_expiry else None,
            S3_STATUS_LABELS.get(r.s3_status, "—") if r.s3_status else "—",
            r.s3_bucket_name,
            r.checked_at.strftime("%d/%m/%Y %H:%M"),
        ]
        ws.append(row)
        line = ws.max_row
        ws.cell(row=line, column=4).fill = STATUS_FILL[r.status]
        exposure_fill = _exposure_fill(r)
        if exposure_fill:
            ws.cell(row=line, column=2).fill = exposure_fill
            ws.cell(row=line, column=2).font = Font(color="FFFFFF", bold=True)
        ws.cell(row=line, column=9).fill = PROTECTED_FILL if r.akamai_protected else UNPROTECTED_FILL
        if r.s3_status == "public":
            ws.cell(row=line, column=14).fill = S3_PUBLIC_FILL
            ws.cell(row=line, column=14).font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws, [32, 14, 40, 10, 12, 20, 16, 16, 14, 40, 40, 24, 14, 14, 32, 18])


def build_xlsx_report(scan_run: ScanRun, results: list[ScanResult], project_name: str = "") -> io.BytesIO:
    summary = compute_scan_summary(results)
    wb = Workbook()
    _build_summary_sheet(wb, scan_run, summary, project_name)
    _build_risk_sheet(wb, results)
    _build_s3_sheet(wb, results)
    _build_detail_sheet(wb, results)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
